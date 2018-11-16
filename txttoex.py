import tkinter
import tkinter.messagebox
import openpyxl

root = tkinter.Tk()
file = open("../logs.txt","r")
wb = openpyxl.load_workbook('testfile.xlsx')
ws = wb.worksheets[0]
#code to add widgets go here

#callback will take whatever is in the file and then add it to the xl file
def callBack():   
    excelFormat = formatList()
    curRow = ws.max_row
    #here is where the code to add data to the excel file lies
    for x in excelFormat:
        for y in x:
            for lenL in range(0,len(y)):
                ws.cell(row=curRow,column=lenL+1).value=y[lenL]
            curRow+=1
    tkinter.messagebox.showinfo("Complete","The task is done.")
    file2 = open("../logs.txt","w").close()
    wb.save("testfile.xlsx")

#formats the read text into a list of lists (easier for me to comprehend)
def formatList():
    smList = []
    finalList = []
    for line in file:
        if line[-1] == "\n":
            line = line[:-1]
        smList.append(line.split(","))
        finalList.append(smList)
        smList = []   
    return finalList

def incAge():
    return

#this button will start the program to add from the textfile to the xl file
b = tkinter.Button(root, text="Convert from text to excel!", command=callBack)

b.pack()
b.place(x=75, y=125)
root.geometry('{}x{}'.format(300,300))
root.mainloop()
